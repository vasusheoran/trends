
import openpyxl
from pathlib import Path
import sys
from scipy.optimize import brentq

# Add trends-py to path
sys.path.insert(0, str(Path.cwd() / "trends-py"))

from app.engine.indicators import EMAState, TOLERANCE
from app.engine.futures import _ce2, _search_br, _search_cc, _bp_series

EXCEL_PATH = Path.cwd().parent / "data" / "Final-Bullish-CE.xlsx"
SHEET_NAME = "Nifty-20.12.2024"

def get_row(row_num):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    
    data = {
        "row": row_num,
        "date": ws[f"B{row_num}"].value,
        "close": ws[f"W{row_num}"].value,
        "ema5": ws[f"AS{row_num}"].value,
        "ema20": ws[f"BN{row_num}"].value,
        "support": ws[f"CC{row_num}"].value,
        "bullish": ws[f"BR{row_num}"].value,
    }
    
    # Also get previous row for ema_pre
    prev_row = row_num - 1
    data["prev"] = {
        "ema5": ws[f"AS{prev_row}"].value,
        "ema20": ws[f"BN{prev_row}"].value,
        "cd": ws[f"BZ{prev_row}"].value, 
    }
    
    wb.close()
    return data

def _get_support(ema5_pre, ema20_pre, cd_pre):
    # First update CD with CE2 of the current bar
    ce2 = _ce2(ema5_pre, ema20_pre)
    cd_ema = EMAState(period=5, decay=2/6, value=cd_pre, seeded=True)
    cd_ema.update(ce2)
    cd_curr = cd_ema.value
    
    # Now find Support
    try:
        cc_trial = brentq(_search_cc, 0, 99999, args=(ema5_pre, ema20_pre, cd_curr))
        support = (2/6) * (cc_trial - cd_curr) + cd_curr
        return support
    except:
        return None

def _search_bullish_new(trial, ema5_pre_today, ema20_pre_today, cd_pre_today):
    # Day 1: Close = trial
    ema5_post_today = ema5_pre_today.copy()
    ema5_post_today.update(trial)
    ema20_post_today = ema20_pre_today.copy()
    ema20_post_today.update(trial)
    
    # cd_post_today = CD updated with CE2(today)
    ce2_today = _ce2(ema5_pre_today, ema20_pre_today)
    cd_ema_today = EMAState(period=5, decay=2/6, value=cd_pre_today, seeded=True)
    cd_ema_today.update(ce2_today)
    cd_post_today = cd_ema_today.value
    
    support_tomorrow = _get_support(ema5_post_today, ema20_post_today, cd_post_today)
    if support_tomorrow is None:
        return 99999.0
    
    return support_tomorrow - trial

def _search_ce_general(trial, ema5_pre, ema20_pre, fixed_closes):
    closes = fixed_closes + [trial, trial]
    bp = _bp_series(ema5_pre, ema20_pre, closes)
    if any(b is None for b in bp):
        return 0.0
    return bp[-1] - bp[-2]

def _search_bullish_verified(trial, ema5_pre, ema20_pre, cd3, close):
    # W1=close, W2=W3=CD4(trial), W4=W5=trial
    # find BP[d+5]=BP[d+4]
    cd4 = (2/6) * (trial - cd3) + cd3
    closes = [close, cd4, cd4, trial, trial]
    bp = _bp_series(ema5_pre, ema20_pre, closes)
    if any(b is None for b in bp):
        return 0.0
    return bp[-1] - bp[-2]

def _get_bullish_verified(ema5_pre, ema20_pre, cd_pre, close):
    # CE2 (today)
    ce2 = _ce2(ema5_pre, ema20_pre)
    cd2 = (2/6) * (ce2 - cd_pre) + cd_pre
    
    # CE3: W2=close, W3=W4=trial, find BP[d+3]=BP[d+2]
    ce3 = brentq(_search_ce_general, 0, 99999, args=(ema5_pre, ema20_pre, [close]))
    
    # CD3
    cd3 = (2/6) * (ce3 - cd2) + cd2
    
    # bullish_trial: W1=close, W2=W3=CD4, W4=W5=trial, find BP[d+5]=BP[d+4]
    bullish_trial = brentq(_search_bullish_verified, 0, 99999, args=(ema5_pre, ema20_pre, cd3, close))
    
    bullish = (2/6) * (bullish_trial - cd3) + cd3
    return bullish

def main():
    for row_num in range(5380, 5388):
        try:
            data = get_row(row_num)
        except:
            continue
        print(f"Row {row_num} ({data['date']})")
        
        ema5_pre = EMAState(period=5, decay=2/6, value=data['prev']['ema5'], seeded=True)
        ema20_pre = EMAState(period=20, decay=2/21, value=data['prev']['ema20'], seeded=True)
        cd_pre = data['prev']['cd']
        
        # Try current Python Bullish (Old)
        ce2 = _ce2(ema5_pre, ema20_pre)
        try:
            bullish_old = brentq(_search_br, 0, 99999, args=(ema5_pre, ema20_pre, ce2))
            print(f"  Old: {bullish_old:.2f} (diff: {bullish_old - data['bullish']:6.2f})")
        except:
            print("  Old: failed")

        # Try New Python Bullish (Support(tomorrow) == trial)
        try:
            bullish_new = brentq(_search_bullish_new, 0, 99999, args=(ema5_pre, ema20_pre, cd_pre))
            print(f"  New: {bullish_new:.2f} (diff: {bullish_new - data['bullish']:6.2f})")
        except:
            print("  New: failed")
            
        # Try Verified Bullish
        try:
            bullish_v = _get_bullish_verified(ema5_pre, ema20_pre, cd_pre, data['close'])
            print(f"  Ver: {bullish_v:.2f} (diff: {bullish_v - data['bullish']:6.2f})")
        except:
            print("  Ver: failed")
            
        print("-" * 35)

if __name__ == "__main__":
    main()
