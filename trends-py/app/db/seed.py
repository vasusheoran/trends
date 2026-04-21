"""
Seed EMA state from Excel or TimescaleDB on startup.

Decision order:
  1. If TimescaleDB has >= 50 rows for the ticker → load from DB.
  2. Otherwise → read from Excel (Final-bullish-ce.xlsx, Nifty-20.12.2024 sheet, rows 5-5387).
"""

import openpyxl
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from app.engine.state import TickerState


_SHEET_NAME = "Nifty-20.12.2024"
_DATA_START = 5
_DATA_END = 5387


def seed_from_excel(state: "TickerState", excel_path: str) -> int:
    """
    Feed historical bars into TickerState from Excel.
    Returns number of bars loaded.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel seed file not found: {excel_path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[_SHEET_NAME]

    count = 0
    for row in range(_DATA_START, _DATA_END + 1):
        date = ws[f"B{row}"].value
        close = ws[f"W{row}"].value
        open_ = ws[f"X{row}"].value
        high = ws[f"Y{row}"].value
        low = ws[f"Z{row}"].value

        if not all([date, close, open_, high, low]):
            continue

        if hasattr(date, "strftime"):
            date = date.strftime("%d-%b-%Y")

        state.update(str(date), float(close), float(open_), float(high), float(low))
        count += 1

    wb.close()
    return count


async def seed_from_db(state: "TickerState", ticker: str) -> int:
    """
    Feed bars from TimescaleDB into TickerState.
    Returns number of bars loaded.
    """
    from app.db.timescale import load_bars
    bars = await load_bars(ticker, limit=101)
    for b in bars:
        state.update(b["date"], b["close"], b["open"], b["high"], b["low"])
    return len(bars)


async def seed_state(state: "TickerState", ticker: str, excel_path: str) -> str:
    """
    Seed state using DB if available, else Excel.
    Returns 'db' or 'excel' indicating which source was used.
    Calls state.commit() after loading to switch to live mode.
    """
    from app.db.timescale import get_row_count
    count = await get_row_count(ticker)
    if count >= 50:
        await seed_from_db(state, ticker)
    else:
        seed_from_excel(state, excel_path)
    state.commit()
    return "db" if count >= 50 else "excel"
