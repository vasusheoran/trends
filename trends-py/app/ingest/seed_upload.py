"""
POST /api/seed/{ticker} — upload an Excel or CSV file to seed historical state.

Excel Column detection: scans every row for a header row containing Date, Close, Open, High, Low
(case-insensitive). Fallback: B=Date, W=Close, X=Open, Y=High, Z=Low.

CSV Column detection: expects headers named Date, Close, Open, High, Low.

Rows where any of those five values are missing or non-numeric are silently skipped.
Resets existing ticker state before loading.
"""

import logging
import openpyxl
import pandas as pd
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from app.registry import reset_state

router = APIRouter()
log = logging.getLogger(__name__)

_REQUIRED = {"date", "close", "open", "high", "low"}
# Fallback fixed column indices (0-based, full row: A=0, B=1, ..., W=22, X=23, Y=24, Z=25)
_FALLBACK = {"date": 1, "close": 22, "open": 23, "high": 24, "low": 25}


def _detect_header(ws) -> Optional[dict]:
    """
    Scan up to the first 20 rows for a header row containing all required column names.
    Returns a dict mapping field name → 0-based column index within the full row (col 1 = index 0).
    Returns None if no header found.
    """
    for row in ws.iter_rows(max_row=20, values_only=True):
        normalized = [str(c).strip().lower() if c is not None else "" for c in row]
        if _REQUIRED.issubset(set(normalized)):
            return {field: normalized.index(field) for field in _REQUIRED}
    return None


@router.post("/api/seed/{ticker}")
async def seed_ticker(
    ticker: str,
    file: UploadFile = File(...),
    sheet: Optional[str] = Form(default=None),
):
    """
    Upload an Excel (.xlsx) or CSV (.csv) file to seed historical OHLCV data for a ticker.

    Optional form field `sheet`: name of the sheet to read (Excel only).
    """
    log.info("Seed request for %s: file=%s", ticker, file.filename)
    contents = await file.read()
    if not contents:
        log.warning("Empty file uploaded for %s", ticker)
        raise HTTPException(status_code=400, detail="File is empty")

    state = reset_state(ticker)
    count = 0
    detection = "unknown"

    fname = file.filename.lower()
    if fname.endswith(".csv") or fname.endswith(".txt"):
        try:
            # Try comma first, then semicolon, then tab
            df = None
            for sep in [',', ';', '\t']:
                try:
                    df = pd.read_csv(BytesIO(contents), skipinitialspace=True, sep=sep)
                    # Normalize column names
                    df.columns = [str(c).strip().lower() for c in df.columns]
                    if _REQUIRED.issubset(set(df.columns)):
                        detection = f"CSV ({sep})"
                        log.info("Detected CSV with delimiter '%s'", sep)
                        break
                    df = None
                except Exception:
                    continue
            
            if df is None:
                log.error("CSV header detection failed for %s. Found columns: %s", ticker, list(df.columns if 'df' in locals() and df is not None else []))
                raise HTTPException(
                    status_code=400, 
                    detail=f"CSV must contain headers: {list(_REQUIRED)}. Check delimiters (comma/semicolon/tab)."
                )
            
            for _, row in df.iterrows():
                try:
                    d, c, o, h, l = row["date"], row["close"], row["open"], row["high"], row["low"]
                    if pd.isna(d) or pd.isna(c): continue
                    state.update(str(d), float(c), float(o), float(h), float(l))
                    count += 1
                except (TypeError, ValueError):
                    continue
        except HTTPException:
            raise
        except Exception as e:
            log.exception("CSV seed error for %s", ticker)
            raise HTTPException(status_code=400, detail=f"Could not parse CSV: {str(e)}")
    
    else:
        # Excel parsing logic
        try:
            wb = openpyxl.load_workbook(BytesIO(contents), data_only=True, read_only=True)
        except Exception as e:
            log.error("Excel load failed for %s: %s", ticker, str(e))
            raise HTTPException(status_code=400, detail="Could not parse file — make sure it is a valid .xlsx or .csv file")

        if sheet is not None:
            if sheet not in wb.sheetnames:
                wb.close()
                raise HTTPException(
                    status_code=400,
                    detail=f"Sheet '{sheet}' not found. Available sheets: {wb.sheetnames}",
                )
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]
        
        col_map = _detect_header(ws)
        using_fallback = col_map is None
        if using_fallback:
            col_map = _FALLBACK
            log.info("Using Excel fallback mapping (B/W/X/Y/Z) for %s", ticker)

        for row in ws.iter_rows(values_only=True):
            if not any(row): continue
            
            try:
                date  = row[col_map["date"]]  if col_map["date"]  < len(row) else None
                close = row[col_map["close"]] if col_map["close"] < len(row) else None
                open_ = row[col_map["open"]]  if col_map["open"]  < len(row) else None
                high  = row[col_map["high"]]  if col_map["high"]  < len(row) else None
                low   = row[col_map["low"]]   if col_map["low"]   < len(row) else None

                if not all([date, close, open_, high, low]):
                    continue

                close = float(close)
                open_ = float(open_)
                high  = float(high)
                low   = float(low)

                if hasattr(date, "strftime"):
                    date = date.strftime("%d-%b-%Y")

                state.update(str(date), close, open_, high, low)
                count += 1
            except (TypeError, ValueError, IndexError):
                continue
        
        detection = "fallback (B/W/X/Y/Z)" if using_fallback else "Excel headers"
        wb.close()

    log.info("Successfully seeded %s: %d bars loaded via %s", ticker, count, detection)

    return {
        "ticker": ticker.upper(),
        "bars_loaded": count,
        "status": "seeded" if count > 0 else "no_data",
        "column_detection": detection,
    }
